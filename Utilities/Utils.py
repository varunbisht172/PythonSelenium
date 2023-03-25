import openpyxl



def readExcelCustom():
        da = []
        wb = openpyxl.load_workbook(filename="C:\\Users\\Manish Bisht\\sampleData.xlsx")
        sh = wb["Sheet1"]
        row_count = sh.max_row
        col_count = sh.max_column
        for i in range(1, row_count):
            r = []
            for j in range(0, col_count):
                r.append(sh.cell(i, j))

            da.append(r)
        return da